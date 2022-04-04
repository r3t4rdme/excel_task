import os
import sqlite3
from sqlite3 import Error

import openpyxl

create_users_table = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    second_name TEXT NOT NULL,
    first_name TEXT NOT NULL,
    patronymic TEXT NOT NULL,
    region_id INTEGER NOT NULL,
    city_id INTEGER NOT NULL,
    phone TEXT NOT NULL,
    email TEXT NOT NULL
    );
INSERT INTO users(
    second_name, first_name, patronymic, region_id, city_id, phone, email)
VALUES
    ('Перов', 'Игорь', 'Алексеевич', 2, 1, '+7 999 200 53 79',
        'perov@yandex.ru'),
    ('Иванов', 'Иван', 'Иванович', 1, 1, '+7 939 200 53 79',
        'ivanov@yandex.ru'),
    ('Григорьев', 'Григорий', 'Григорьевич', 2, 2, '+7 949 200 53 79',
        'grigoriev@yandex.ru')
    """

create_regions_table = """
CREATE TABLE IF NOT EXISTS regions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    region_name TEXT NOT NULL
    );
INSERT INTO regions(region_name)
VALUES
    ('Краснодарский край'),
    ('Ростовская область'),
    ('Ставропольский край')
    """

create_cities_table = """
CREATE TABLE IF NOT EXISTS cities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    region_id INTEGER NOT NULL,
    city_name TEXT NOT NULL,
    FOREIGN KEY (region_id) REFERENCES regions (id)
    );
INSERT INTO cities(region_id, city_name)
VALUES
    (0, 'Краснодар'),
    (0, 'Кропоткин'),
    (0, 'Славянск'),
    (1, 'Ростов'),
    (1, 'Шахты'),
    (1, 'Батайск'),
    (2, 'Ставрополь'),
    (2, 'Пятигорск'),
    (2, 'Кисловодск')
    """


def main():

    def create_connection(path):
        connection = None
        try:
            if os.path.isfile(path):
                connection = sqlite3.connect(path)
                print('Соединение с БД успешно установлено')
            else:
                connection = sqlite3.connect(path)
                cursor = connection.cursor()
                try:
                    cursor.executescript(create_regions_table)
                    cursor.executescript(create_cities_table)
                    cursor.executescript(create_users_table)
                except Error as e:
                    print(f"Ошибка во время заполнения БД: {e}")

        except Error as e:
            print(f"Ошибка во время заполнения БД: {e}")

        return connection

    def import_excel(input_path, connection, min_row, max_row, max_col):
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        connection = connection
        cursor = connection.cursor()
        import_query = """
            INSERT INTO users(
            second_name, first_name, patronymic,
            region_id, city_id, phone, email)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """
        try:
            for row in ws.iter_rows(
                    min_row=min_row, max_row=max_row, max_col=max_col):
                data = [cell.value for cell in row]
                cursor.execute(import_query, data)
        except Exception:
            print('Проблема с импортом данных из таблицы ')

        # не самый оптимальный метод сменить id-шники,
        # но по другому пока не придумал..
        update_region_id_query = """
        UPDATE users
        SET
            region_id = (SELECT regions.id
                            FROM regions
                            WHERE regions.region_name = users.region_id)
        WHERE
            EXISTS (
                SELECT *
                FROM regions
                WHERE regions.region_name = users.region_id
            )
        """
        update_city_id_query = """
        UPDATE users
        SET
            city_id = (SELECT cities.id
                            FROM cities
                            WHERE cities.city_name = users.city_id)
        WHERE
            EXISTS (
                SELECT *
                FROM cities
                WHERE cities.city_name = users.city_id
            )
        """
        try:
            cursor.execute(update_region_id_query)
            cursor.execute(update_city_id_query)
        except Exception:
            print('Проблемы с загрузкой данных в БД ')

        connection.commit()
        connection.close

    def export_excel(connection, filename):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(title='users', index=0)
        cursor = connection.cursor()
        # получаем ID для users из regions и cities
        query = cursor.execute(
            """SELECT second_name, first_name, patronymic,
            region_name, city_name, phone, email
            FROM users, regions, cities
            WHERE users.region_id = regions.id
            AND users.city_id = cities.id""")
        # собираем названия колонок
        column_names = [member[0] for member in query.description]
        # импортируем названия колонок в файл
        ws.append(column_names)
        # заполняем файл данными из таблицы
        for row in query:
            ws.append(row)
        # сохраняем экспортный файл под названием полученным от пользователя
        wb.save(filename=filename)

    connection_block = str(input(
        'Введите название новой или путь к старой БД: '))
    connection = create_connection(connection_block)

    while True:

        try:
            start_block = int(input(
                'Что необходимо сделать?\n '
                '1 - Импорт из файла .xlsx\n '
                '2 - Экспорт в файл .xlsx\n '
                '0 - Выход '
            ))

            if start_block == 1:
                connection = connection

                try:

                    input_path = input('Введите путь к файлу ') + '.xlsx'
                    min_row = int(input(
                        'Введите начальную строку нужной таблицы '))
                    max_row = int(input(
                        'Введите конечную строку нужной таблицы '))
                    max_col = int(input(
                        'Введите крайнюю колонку нужной таблицы '))
                    import_excel(
                        input_path=input_path, connection=connection,
                        min_row=min_row, max_row=max_row, max_col=max_col)

                except Exception:
                    print('Указан неверный путь к файлу ')

            elif start_block == 2:
                connection = connection
                filename = str(input(
                    'Введите название экспортируемого файла ') + '.xlsx')
                export_excel(connection=connection, filename=filename)

            elif start_block == 0:
                exit()
        except Exception:
            print('Выберите нужное действие ')


if __name__ == '__main__':
    main()
